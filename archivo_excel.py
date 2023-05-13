from openpyxl import load_workbook

""""
Para manejar archivos Excel en Python, puedes utilizar la librería "openpyxl", que permite trabajar con archivos de formato xlsx. A continuación, te proporciono algunos ejemplos básicos de cómo manejar archivos Excel en Python utilizando la librería openpyxl:"""

# Abrir archivo
wb = load_workbook('archivo.xlsx')

# Seleccionar hoja de trabajo
ws = wb.active

# Leer valor de una celda
value = ws['A1'].value

# Leer valores de un rango de celdas
for row in ws['A1':'C3']:
    for cell in row:
        print(cell.value)

# Escribir valor en una celda
ws['A1'] = 'Valor'

# Escribir valores en un rango de celdas
values = [
    ['Valor1', 'Valor2', 'Valor3'],
    ['Valor4', 'Valor5', 'Valor6'],
]
for row in range(len(values)):
    for col in range(len(values[row])):
        ws.cell(row=row+1, column=col+1, value=values[row][col])

# Guardar cambios en el archivo
wb.save('archivo.xlsx')